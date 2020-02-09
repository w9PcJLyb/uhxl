import pandas as pd
from pandas.io.excel._util import _maybe_convert_usecols
from pandas.io.excel._openpyxl import _OpenpyxlReader
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet


class UhExcelFile(pd.ExcelFile):
    """
    Class for parsing tabular excel sheets into DataFrame objects,
    allows to read only visible sheets, columns and rows.

    Parameters:
    ----------
    io: str, path or file-like (see pandas.ExcelFile)
    hide_sheets: if True -> read only visible sheets
    hide_columns: if True -> read only visible columns
    hide_rows: if True -> read only visible rows

    Example:
    ----------
    >>> excel_file = UhExcelFile("path/to/file.xlsx")
    >>> df = pd.read_excel(excel_file)
    """

    def __init__(self, io, hide_sheets=True, hide_columns=True, hide_rows=True):
        super().__init__(io, engine="openpyxl")
        # ignore engine, redefine reader
        self._reader = UhOpenpyxlReader(
            self._io,
            hide_sheets=hide_sheets,
            hide_columns=hide_columns,
            hide_rows=hide_rows,
        )


class UhOpenpyxlReader(_OpenpyxlReader):
    def __init__(
        self, filepath_or_buffer, hide_sheets=True, hide_columns=True, hide_rows=True
    ):
        super().__init__(filepath_or_buffer)
        self.hide_sheets = hide_sheets
        self.hide_columns = hide_columns
        self.hide_rows = hide_rows

        # forbid save
        self.book._read_only = True

        # parse arguments
        self.usecols = None
        self.skiprows = None

    def load_workbook(self, filepath_or_buffer):
        return load_workbook(
            filepath_or_buffer,
            read_only=False,  # need False to get sheet_state information
            data_only=True,
            keep_links=False,
        )

    @property
    def all_sheets(self):
        return self.book.worksheets

    @property
    def hidden_sheets(self):
        return [
            s for s in self.all_sheets if s.sheet_state != Worksheet.SHEETSTATE_VISIBLE
        ]

    @property
    def visible_sheets(self):
        return [
            s for s in self.all_sheets if s.sheet_state == Worksheet.SHEETSTATE_VISIBLE
        ]

    @property
    def sheet_names(self):
        if self.hide_sheets:
            return [s.title for s in self.visible_sheets]
        else:
            return [s.title for s in self.all_sheets]

    def get_sheet_by_name(self, name):
        if self.hide_sheets:
            if name in [s.title for s in self.hidden_sheets]:
                raise KeyError(
                    "Worksheet '{}' exist but hidden. Set skip_hidden_sheets=True to get this sheet.".format(
                        name
                    )
                )

        return self.book[name]

    def get_sheet_by_index(self, index):
        if self.hide_sheets:
            return self.visible_sheets[index]
        else:
            return self.all_sheets[index]

    def hidden_rows(self, sheet):
        rows = set()
        if not self.hide_rows:
            return rows

        for let, dim in sheet.row_dimensions.items():
            if dim.hidden:
                rows.add(let)
        return rows

    def hidden_columns(self, sheet):
        columns = set()
        if not self.hide_columns:
            return columns

        for dim in sheet.column_dimensions.values():
            if dim.hidden:
                for col in range(dim.min, dim.max + 1):
                    columns.add(col)
        return columns

    def get_visible_cells(self, sheet):
        hidden_rows = self.hidden_rows(sheet)
        hidden_columns = self.hidden_columns(sheet)
        usecols = self.usecols
        skiprows = self.skiprows

        sheet_cells = sheet._cells
        if (
            not hidden_rows
            and not hidden_columns
            and usecols is None
            and skiprows is None
        ):
            return sheet_cells

        row_filter = {}
        offset = 1
        for row in range(sheet.min_row, sheet.max_row + 1):
            if row not in hidden_rows and (skiprows is None or row not in skiprows):
                row_filter[row] = offset
                offset += 1

        col_filter = {}
        offset = 1
        for col in range(sheet.min_column, sheet.max_column + 1):
            if col not in hidden_columns and (usecols is None or col in usecols):
                col_filter[col] = offset
                offset += 1

        visible_cells = {}
        for (row, col), cell in sheet_cells.items():
            new_row = row_filter.get(row)
            new_col = col_filter.get(col)

            # skip hidden
            if new_row is None or new_col is None:
                continue

            # skip empty cells
            # warning: it change openpyxl engine behavior
            if cell.value is None or cell.data_type == "e":
                continue

            visible_cells[(new_row, new_col)] = cell

            # rebase
            cell.row = new_row
            cell.column = new_col

        return visible_cells

    def get_sheet_data(self, sheet, convert_float):
        sheet_cells = sheet._cells
        sheet._cells = self.get_visible_cells(sheet)
        data = super().get_sheet_data(sheet, convert_float)
        sheet._cells = sheet_cells
        return data

    def parse(self, *args, **kwargs):
        self.usecols = None
        self.skiprows = None

        usecols = kwargs.pop("usecols")
        if usecols:
            usecols = _maybe_convert_usecols(usecols)
            self.usecols = set(c + 1 for c in usecols)

        skiprows = kwargs.pop("skiprows")
        if skiprows:
            self.skiprows = set(c + 1 for c in skiprows)

        return super().parse(*args, **kwargs)
